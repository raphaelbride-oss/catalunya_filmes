#!/usr/bin/env python3
"""Parse the Catalunya XLSX directly. Replaces the previous markdown-dump
based parser, which only saw a fraction of the rows.

Reads /tmp/planilha.xlsx (downloaded via Drive API) and produces data.json.

Sheets used:
    - CONTAS A PAGAR         -> ap_main          (~2 500 rows)
    - Renegociados            -> ap_reneg         (~1 100 rows, header on row 4)
    - AR                      -> ar               (~160 rows)
    - Cash Flow Daily         -> saldo_diario     (rolling balance per day)
"""
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

XLSX = Path("/tmp/planilha.xlsx")
OUT = Path("/Users/raphabride/Desktop/fluxo-caixa-dash/data.json")


def to_iso(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        if v.year < 2000 or v.year > 2100:
            return None
        return v.date().isoformat()
    if isinstance(v, date):
        if v.year < 2000 or v.year > 2100:
            return None
        return v.isoformat()
    if isinstance(v, str):
        t = v.strip()
        if not t:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                d = datetime.strptime(t, fmt).date()
                if 2000 <= d.year <= 2100:
                    return d.isoformat()
            except ValueError:
                continue
    return None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN
            return None
        return float(v)
    if isinstance(v, str):
        t = v.strip().replace("R$", "").replace(" ", "")
        if not t:
            return None
        neg = False
        if t.startswith("(") and t.endswith(")"):
            neg = True
            t = t[1:-1]
        if t.startswith("-"):
            neg = True
            t = t[1:]
        t = t.replace(".", "").replace(",", ".")
        try:
            f = float(t)
            return -f if neg else f
        except ValueError:
            return None
    return None


def s(v):
    if v is None:
        return ""
    return str(v).strip()


def find_header_row(rows, must_contain):
    """Return index of the first row that contains all `must_contain` strings (case-insensitive)."""
    keys = [k.lower() for k in must_contain]
    for i, r in enumerate(rows):
        cells = [s(c).lower() for c in r if c is not None]
        joined = " | ".join(cells)
        if all(k in joined for k in keys):
            return i
    return None


def build_index(header_row):
    return {s(h).upper().replace(" ", " ").strip(): i for i, h in enumerate(header_row) if h is not None}


def col(idx, name, row):
    """Get cell value by header name (case-insensitive, ignoring trailing spaces)."""
    name_up = name.upper().strip()
    for k, v in idx.items():
        if k.replace(" ", "").upper() == name_up.replace(" ", "").upper():
            return row[v] if v < len(row) else None
    return None


wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)


# --------------------------------------------------------------------------
# CONTAS A PAGAR (AP principal)
# --------------------------------------------------------------------------
ap_main = []
ws = wb["CONTAS A PAGAR"]
rows = list(ws.iter_rows(values_only=True))
hdr_idx = find_header_row(rows, ["COMPETENCIA", "FORNECEDOR", "VENCIMENTO ORIGINAL"])
if hdr_idx is None:
    raise SystemExit("Header de CONTAS A PAGAR não encontrado")
header = rows[hdr_idx]
idx = build_index(header)

for r in rows[hdr_idx + 1:]:
    # ignora linhas inteiramente vazias
    if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
        continue
    fornecedor = s(col(idx, "FORNECEDOR", r))
    valor = to_float(col(idx, "VALOR FACE", r)) or to_float(col(idx, "VALOR A PAGAR", r))
    if not fornecedor and valor is None:
        continue
    ap_main.append({
        "competencia": to_iso(col(idx, "COMPETENCIA", r)) or s(col(idx, "COMPETENCIA", r)),
        "empresa": s(col(idx, "EMPRESA", r)),
        "tipo_conta": s(col(idx, "TIPO DE CONTA", r)),
        "item_cash": s(col(idx, "Item Cash", r)),
        "fornecedor": fornecedor,
        "vencimento": to_iso(col(idx, "VENCIMENTO ORIGINAL", r)),
        "previsao": to_iso(col(idx, "PREVISÃO PAGAMENTO", r)),
        "pagamento": to_iso(col(idx, "DATA DE PAGAMENTO", r)),
        "valor": valor,
        "n_job": s(col(idx, "Nº JOB", r)),
        "job": s(col(idx, "JOB", r)),
        "descricao": s(col(idx, "DESCRIÇÃO", r)),
        "forma_pagamento": s(col(idx, "FORMA DE PAGAMENTO", r)),
        "plano_contas": "",  # AP main não tem essa coluna
        "status": s(col(idx, "STATUS", r)),
        "dias_atraso": to_float(col(idx, "DIAS EM ATRASO", r)),
    })


# --------------------------------------------------------------------------
# Renegociados (AP renegociado) — header na linha 4 (índice 3)
# --------------------------------------------------------------------------
ap_reneg = []
ws = wb["Renegociados"]
rows = list(ws.iter_rows(values_only=True))
hdr_idx = find_header_row(rows, ["FORNECEDOR", "VENCIMENTO ORIGINAL", "PREVISÃO PAGAMENTO"])
if hdr_idx is None:
    raise SystemExit("Header de Renegociados não encontrado")
header = rows[hdr_idx]
idx_reneg = build_index(header)

for r in rows[hdr_idx + 1:]:
    if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
        continue
    fornecedor = s(col(idx_reneg, "FORNECEDOR", r))
    valor = to_float(col(idx_reneg, "VALOR", r)) or to_float(col(idx_reneg, "VALOR FACE", r))
    if not fornecedor and valor is None:
        continue
    ap_reneg.append({
        "status_reneg": s(col(idx_reneg, "STATUS", r)),
        "empresa": s(col(idx_reneg, "EMPRESA", r)),
        "tipo_conta": s(col(idx_reneg, "TIPO DE CONTA", r)),
        "item_cash": s(col(idx_reneg, "Item Cash", r)),
        "fornecedor": fornecedor,
        "vencimento": to_iso(col(idx_reneg, "VENCIMENTO ORIGINAL", r)),
        "previsao": to_iso(col(idx_reneg, "PREVISÃO PAGAMENTO", r)),
        "pagamento": to_iso(col(idx_reneg, "DATA DE PAGAMENTO", r)),
        "competencia": to_iso(col(idx_reneg, "COMPETÊNCIA", r)) or s(col(idx_reneg, "COMPETÊNCIA", r)),
        "valor": valor,
        "parcelas": s(col(idx_reneg, "Parcelas", r)),
        "n_job": s(col(idx_reneg, "Nº JOB", r)),
        "job": s(col(idx_reneg, "JOB", r)),
        "descricao": s(col(idx_reneg, "DESCRIÇÃO", r)),
        "forma_pagamento": s(col(idx_reneg, "FORMA DE PAGAMENTO", r)),
        "plano_contas": s(col(idx_reneg, "PLANO DE CONTAS", r)),
        "status": "",  # derivado abaixo
        "dias_atraso": to_float(col(idx_reneg, "DIAS EM ATRASO", r)),
    })


# --------------------------------------------------------------------------
# AR
# --------------------------------------------------------------------------
ar = []
ws = wb["AR"]
rows = list(ws.iter_rows(values_only=True))
hdr_idx = find_header_row(rows, ["CLIENTE", "PREVISÃO RECEBIMENTO"])
if hdr_idx is None:
    raise SystemExit("Header de AR não encontrado")
header = rows[hdr_idx]
idx_ar = build_index(header)

for r in rows[hdr_idx + 1:]:
    if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
        continue
    cliente = s(col(idx_ar, "CLIENTE", r))
    valor = to_float(col(idx_ar, "VALOR", r))
    if not cliente and valor is None:
        continue
    ar.append({
        "empresa": s(col(idx_ar, "EMPRESA", r)),
        "tipo_conta": s(col(idx_ar, "TIPO DE CONTA", r)),
        "vencimento": to_iso(col(idx_ar, "VENCIMENTO ORIGINAL", r)),
        "previsao": to_iso(col(idx_ar, "PREVISÃO RECEBIMENTO", r)),
        "recebimento": to_iso(col(idx_ar, "DATA DE RECEBIMENTO", r)),
        "competencia": to_iso(col(idx_ar, "COMPETÊNCIA", r)) or s(col(idx_ar, "COMPETÊNCIA", r)),
        "cliente": cliente,
        "valor": valor,
        "n_job": s(col(idx_ar, "Nº JOB", r)),
        "job": s(col(idx_ar, "JOB", r)),
        "descricao": s(col(idx_ar, "DESCRIÇÃO", r)),
        "forma_pagamento": s(col(idx_ar, "FORMA DE PAGAMENTO", r)),
        "plano_contas": s(col(idx_ar, "PLANO DE CONTAS", r)),
        "status": s(col(idx_ar, "STATUS", r)),
        "dias_atraso": to_float(col(idx_ar, "DIAS EM ATRASO", r)),
    })


# --------------------------------------------------------------------------
# Cash Flow Daily — saldo diário
# Layout: linha de datas + linha de saldos abaixo. Vamos varrer e encontrar.
# --------------------------------------------------------------------------
saldo_diario = []
ws = wb["Cash Flow Daily"]
all_rows = list(ws.iter_rows(values_only=True))

# Procura a primeira linha onde >50 células são datas
date_row_idx = None
for i, r in enumerate(all_rows):
    n = sum(1 for c in r if isinstance(c, (datetime, date)) and not isinstance(c, datetime) or isinstance(c, datetime))
    if n > 50:
        date_row_idx = i
        break

if date_row_idx is not None:
    date_row = all_rows[date_row_idx]
    # A linha de valores normalmente é a próxima com números
    for j in range(date_row_idx + 1, min(date_row_idx + 8, len(all_rows))):
        val_row = all_rows[j]
        n_nums = sum(1 for c in val_row if isinstance(c, (int, float)))
        if n_nums >= 20:
            for col_i, dt in enumerate(date_row):
                if isinstance(dt, (datetime, date)) and col_i < len(val_row):
                    v = val_row[col_i]
                    if isinstance(v, (int, float)):
                        saldo_diario.append({"date": to_iso(dt), "saldo": float(v)})
            break


# --------------------------------------------------------------------------
# Normaliza empresa e deriva situação
# --------------------------------------------------------------------------
TODAY = date.today()


def norm_empresa(s):
    if not s:
        return s
    return s.strip()


for arr in (ap_main, ap_reneg, ar):
    for row in arr:
        row["empresa"] = norm_empresa(row.get("empresa", ""))


def classify_ap(row):
    if row.get("pagamento"):
        return "Pago"
    p = row.get("previsao") or row.get("vencimento")
    if p:
        try:
            if datetime.fromisoformat(p).date() < TODAY:
                return "Atrasado"
        except (ValueError, TypeError):
            pass
    return "Pendente"


def classify_ar(row):
    if row.get("recebimento"):
        return "Recebido"
    p = row.get("previsao") or row.get("vencimento")
    if p:
        try:
            if datetime.fromisoformat(p).date() < TODAY:
                return "Atrasado"
        except (ValueError, TypeError):
            pass
    return "Pendente"


for r in ap_main:
    r["situacao"] = classify_ap(r)
for r in ap_reneg:
    r["situacao"] = classify_ap(r)
for r in ar:
    r["situacao"] = classify_ar(r)


out = {
    "generated_at": datetime.now().isoformat(timespec="seconds"),
    "today": TODAY.isoformat(),
    "saldo_diario": saldo_diario,
    "ap_main": ap_main,
    "ap_reneg": ap_reneg,
    "ar": ar,
    "parcelas_reneg": [],  # mantido para compat — não usado mais
}

OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2))


# --------------------------------------------------------------------------
# Resumo
# --------------------------------------------------------------------------
def cnt(arr, key):
    out = {}
    for r in arr:
        v = r.get(key) or ""
        out[v] = out.get(v, 0) + 1
    return out


def total(arr):
    return sum((r.get("valor") or 0) for r in arr)


print(f"CONTAS A PAGAR: {len(ap_main):5d} linhas | total R$ {total(ap_main):>15,.2f} | situacao={cnt(ap_main, 'situacao')}")
print(f"Renegociados:   {len(ap_reneg):5d} linhas | total R$ {total(ap_reneg):>15,.2f} | situacao={cnt(ap_reneg, 'situacao')}")
print(f"AR:             {len(ar):5d} linhas | total R$ {total(ar):>15,.2f} | situacao={cnt(ar, 'situacao')}")
print(f"Saldo diario:   {len(saldo_diario):5d} pontos | range={saldo_diario[0]['date']}..{saldo_diario[-1]['date']}" if saldo_diario else "Saldo diario vazio")

# Pendentes em maio/26 — sanity check
mai = [r for r in ap_main if not r.get("pagamento") and r.get("previsao") and r["previsao"].startswith("2026-05")]
mai_total = sum(r["valor"] or 0 for r in mai)
print(f"\n>>> AP a pagar em maio/26 (s/ pagamento, previsao 2026-05): {len(mai)} linhas, total R$ {mai_total:,.2f}")
print(f"Saida: {OUT}")
